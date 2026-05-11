import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook

MONTH_COLUMNS = [str(202600 + month) for month in range(1, 13)]
OUTPUT_HEADERS = [
    "year",
    "category",
    "budget_item",
    "annual_budget",
    "month_01",
    "month_02",
    "month_03",
    "month_04",
    "month_05",
    "month_06",
    "month_07",
    "month_08",
    "month_09",
    "month_10",
    "month_11",
    "month_12",
    "is_valid_expense_item",
    "notes",
]


def normalize_amount(value):
    if value is None or value == "-":
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def load_budget_rows(workbook_path):
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet = workbook["2026Budget"]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {str(value): index for index, value in enumerate(header_row) if value is not None}

    required = ["Year", "Category", "item", "Annual_Budget", "備註(發生費用與期間)"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    month_indexes = []
    for month in MONTH_COLUMNS:
        if month not in headers:
            raise ValueError(f"Missing month column: {month}")
        month_indexes.append(headers[month])

    rows = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        item = row[headers["item"]]
        if not item or str(item).strip() == "預算總額":
            continue

        output = {
            "year": row[headers["Year"]],
            "category": row[headers["Category"]],
            "budget_item": item,
            "annual_budget": normalize_amount(row[headers["Annual_Budget"]]),
            "is_valid_expense_item": "TRUE",
            "notes": row[headers["備註(發生費用與期間)"]] or "",
        }
        for index, month_index in enumerate(month_indexes, start=1):
            output[f"month_{index:02d}"] = normalize_amount(row[month_index])
        rows.append(output)
    return rows


def write_csv(rows, output_path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def main():
    parser = argparse.ArgumentParser(description="Export 2026 budget workbook items to BudgetItems CSV.")
    parser.add_argument("--input", required=True, help="Path to 2026預算表.xlsx")
    parser.add_argument("--output", default="exports/budget_items_import.csv", help="Output CSV path")
    args = parser.parse_args()

    rows = load_budget_rows(Path(args.input))
    write_csv(rows, Path(args.output))
    print(f"Exported {len(rows)} budget item rows to {args.output}")


if __name__ == "__main__":
    main()
