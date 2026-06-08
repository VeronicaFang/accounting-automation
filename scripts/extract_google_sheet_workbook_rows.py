import json
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


def cell_value(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def extract_workbook_rows(input_path):
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    result = {}

    for worksheet in workbook.worksheets:
        row_iterator = worksheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(row_iterator)]
        rows = []

        for row in row_iterator:
            if not any(value is not None and str(value).strip() != "" for value in row):
                continue

            record = {}
            for header, value in zip(headers, row):
                if header:
                    record[header] = cell_value(value)
            rows.append(record)

        result[worksheet.title] = rows

    return result


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: extract_google_sheet_workbook_rows.py <input.xlsx> <output.json>")

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(extract_workbook_rows(input_path), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
